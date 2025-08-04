from flask import Flask, render_template, redirect, url_for, flash, request, session, jsonify, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_migrate import Migrate
from flask_login import LoginManager, login_user, logout_user, login_required, current_user, UserMixin
from werkzeug.security import generate_password_hash, check_password_hash
from flask_wtf.csrf import CSRFProtect, validate_csrf
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
import bleach
import os
from dotenv import load_dotenv
from simple_cache import simple_cache
from performance_monitor import monitor_performance, time_function
from forms import LoginForm, AnimeForm, EpisodeForm, UserForm, EditUserForm, GenreForm, AnimeSearchForm, RegistrationForm, NewsForm, EventForm, CommunityRegistrationForm, CommunityMemberSearchForm, CategoryForm, DeleteForm
from models import db, User, Anime, Episode, Log, Genre, Rating, Notification, News, Event, CommunityInfo, CommunityMember, anime_genres
from community import community_bp, ForumThread
import re
import random
from functools import wraps
import requests
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import io

app = Flask(__name__)
application = app

# Load environment variables
load_dotenv()

# Güvenli konfigürasyon
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'dev-key-change-in-production')
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get('SQLALCHEMY_DATABASE_URI', 'sqlite:///anime_site.db')
app.config['SQLALCHEMY_BINDS'] = {
    'community': os.environ.get('SQLALCHEMY_BINDS_COMMUNITY', 'sqlite:///community.db')
}

# Cache Configuration - Using simple_cache
# app.config['CACHE_TYPE'] = 'simple'
# app.config['CACHE_DEFAULT_TIMEOUT'] = 300

# Session Configuration - Environment based güvenlik ayarları
is_production = not os.environ.get('FLASK_DEBUG', 'False').lower() == 'true'
app.config['SESSION_COOKIE_SECURE'] = os.environ.get('SESSION_COOKIE_SECURE', str(is_production)).lower() == 'true'
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Strict'  # CSRF koruması
app.config['PERMANENT_SESSION_LIFETIME'] = 900  # 15 dakika
app.config['REMEMBER_COOKIE_DURATION'] = 900  # 15 dakika
app.config['REMEMBER_COOKIE_SECURE'] = os.environ.get('REMEMBER_COOKIE_SECURE', str(is_production)).lower() == 'true'
app.config['REMEMBER_COOKIE_HTTPONLY'] = True

# Database optimizations
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    'pool_pre_ping': True,
    'pool_recycle': 300,
    'connect_args': {'check_same_thread': False}
}

db.init_app(app)
migrate = Migrate(app, db)

# CSRF Protection
csrf = CSRFProtect(app)

# Rate Limiting
limiter = Limiter(
    app=app,
    key_func=get_remote_address,
    default_limits=["200 per day", "50 per hour"],
    storage_uri="memory://"
)

# Initialize performance monitoring
monitor_performance(app)

# Register the community blueprint
app.register_blueprint(community_bp, url_prefix='/community/forum')

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Bu sayfaya erişmek için giriş yapmanız gerekiyor.'
login_manager.login_message_category = 'info'
login_manager.session_protection = 'strong'  # Güvenli session koruması

SPECIAL_GENRES = ["Editörün Seçimi", "Hero Section"]

with app.app_context():
    for genre_name in SPECIAL_GENRES:
        if not Genre.query.filter_by(name=genre_name).first():
            db.session.add(Genre(name=genre_name))
    db.session.commit()

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or not (current_user.can_add_user or current_user.can_edit or current_user.can_delete or current_user.is_admin):
            flash('Bu sayfayı görüntüleme yetkiniz yok.', 'danger')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

def strict_admin_required(f):
    """Sadece is_admin=True olan kullanıcılar için - Güçlendirilmiş güvenlik"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            flash('Bu sayfaya erişmek için giriş yapmanız gerekiyor.', 'danger')
            return redirect(url_for('login'))
        
        if not current_user.is_admin:
            # Güvenlik logu
            app.logger.warning(f'🚨 Yetkisiz admin erişim denemesi: Kullanıcı={current_user.username}, IP={request.remote_addr}, URL={request.url}')
            log_action('unauthorized_admin_access', f'Yetkisiz admin erişim denemesi: {request.url}')
            flash('Bu sayfaya sadece adminler erişebilir.', 'danger')
            return redirect(url_for('index'))
        
        # Admin işlem logu
        app.logger.info(f'🔐 Admin erişimi: Kullanıcı={current_user.username}, IP={request.remote_addr}, URL={request.url}')
        return f(*args, **kwargs)
    return decorated_function

def delete_permission_required(f):
    """Silme işlemleri için özel yetki kontrolü"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or not (current_user.can_delete or current_user.is_admin):
            flash('Bu işlemi gerçekleştirme yetkiniz yok.', 'danger')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

def community_access_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or not (current_user.is_admin or current_user.is_community_member):
            flash('Bu sayfaya erişim için admin veya topluluk üyesi olmanız gerekiyor.', 'danger')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

def calendar_access_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or not (current_user.is_admin or current_user.is_community_member):
            flash('Takvim erişimi için admin veya topluluk üyesi olmanız gerekiyor.', 'danger')
            return redirect(url_for('community.index'))
        return f(*args, **kwargs)
    return decorated_function

def update_mal_data(anime):
    try:
        response = requests.get(f'https://api.jikan.moe/v4/anime?q={anime.name}&limit=1')
        if response.status_code == 200:
            data = response.json().get('data')
            if data:
                mal_anime = data[0]
                anime.mal_id = mal_anime.get('mal_id')
                anime.mal_score = mal_anime.get('score')
                anime.mal_url = mal_anime.get('url')
    except requests.exceptions.RequestException as e:
        flash(f'MyAnimeList verileri alınamadı: {e}', 'warning')

def sanitize_input(text):
    """Kullanıcı girdilerini temizle ve güvenli hale getir"""
    if not text:
        return text
    # HTML taglerini temizle
    cleaned = bleach.clean(text, tags=[], attributes={}, strip=True)
    # SQL injection için tehlikeli karakterleri temizle
    cleaned = re.sub(r'[<>"\';\\]|--', '', cleaned)
    return cleaned.strip()

def safe_search_filter(query_text, model_class):
    """SQL Injection'a karşı güvenli arama filtresi"""
    if not query_text:
        return None
    
    # Sadece alfanumerik karakterler, boşluk ve Türkçe karakterler
    safe_query = re.sub(r'[^a-zA-ZğüşıöçĞÜŞİÖÇ0-9\s]', '', query_text)
    safe_query = safe_query.strip()
    
    if not safe_query:
        return None
    
    if model_class == CommunityMember:
        return (
            model_class.name.ilike(f'%{safe_query}%') |
            model_class.surname.ilike(f'%{safe_query}%') |
            model_class.faculty.ilike(f'%{safe_query}%') |
            model_class.department.ilike(f'%{safe_query}%') |
            model_class.student_id.ilike(f'%{safe_query}%')
        )
    elif model_class == Anime:
        return model_class.name.ilike(f'%{safe_query}%')
    
    return None

def log_action(action, description):
    if current_user.is_authenticated:
        new_log = Log(action=action, description=description, user_id=current_user.id)
        db.session.add(new_log)
        db.session.commit()

@login_manager.user_loader
def load_user(user_id):
    try:
        return db.session.get(User, int(user_id))
    except (ValueError, TypeError):
        return None

# Session güvenliği için ek kontrol
@app.before_request
def before_request():
    # Login, logout, register ve statik dosyalar için kontrol yapma
    excluded_endpoints = ['login', 'logout', 'register', 'landing', 'static', 'community_register']
    
    if request.endpoint in excluded_endpoints:
        return
    
    # Session güvenliği kontrolleri
    if current_user.is_authenticated:
        # User'ın hala mevcut olduğunu kontrol et
        user = db.session.get(User, current_user.id)
        if not user:
            logout_user()
            session.clear()
            flash('Oturum süreniz dolmuş. Lütfen tekrar giriş yapın.', 'warning')
            return redirect(url_for('login'))

@app.route('/register', methods=['GET', 'POST'])
@limiter.limit("3 per minute")  # Spam kayıt koruması
def register():
    if current_user.is_authenticated:
        return redirect(url_for('landing'))
    form = RegistrationForm()
    if form.validate_on_submit():
        try:
            # Mükerrer kayıt kontrolü (ek güvenlik)
            existing_username = User.query.filter_by(username=form.username.data).first()
            existing_email = User.query.filter_by(email=form.email.data).first()
            
            if existing_username:
                flash('Bu kullanıcı adı zaten kullanılıyor.', 'error')
                return render_template('register.html', title='Kayıt Ol', form=form)
                
            if existing_email:
                flash('Bu e-posta adresi zaten kayıtlı.', 'error')
                return render_template('register.html', title='Kayıt Ol', form=form)
            
            hashed_password = generate_password_hash(form.password.data, method='pbkdf2:sha256')
            user = User(
                username=form.username.data.strip().lower(),
                email=form.email.data.strip().lower(),
                password=hashed_password
            )
            db.session.add(user)
            db.session.commit()
            
            # Otomatik login yap
            login_user(user, remember=True)
            session.permanent = True
            
            flash(f'Hesabınız başarıyla oluşturuldu! Hoş geldin, {user.username}!', 'success')
            return redirect(url_for('landing'))
            
        except Exception as e:
            db.session.rollback()
            flash('Hesap oluşturulurken bir hata oluştu. Lütfen tekrar deneyin.', 'error')
            app.logger.error(f"User registration error: {str(e)}")
            
    elif request.method == 'POST':
        # Form validasyon hatalarını kullanıcıya göster
        for field, errors in form.errors.items():
            for error in errors:
                flash(f'{getattr(form, field).label.text}: {error}', 'error')
    
    return render_template('register.html', title='Kayıt Ol', form=form)

@app.route('/community/register', methods=['GET', 'POST'])
@limiter.limit("2 per minute")  # Topluluk kayıt spam koruması
def community_register():
    if current_user.is_authenticated:
        return redirect(url_for('landing'))
    form = CommunityRegistrationForm()
    if form.validate_on_submit():
        try:
            # Seçilen birimleri JSON string olarak kaydet
            preferred_units = ','.join(form.preferred_units.data) if form.preferred_units.data else ''
            
            # Mükerrer kayıt kontrolü (ek güvenlik)
            existing_email = CommunityMember.query.filter_by(email=form.email.data).first()
            existing_student_id = CommunityMember.query.filter_by(student_id=form.student_id.data).first()
            existing_phone = CommunityMember.query.filter_by(phone_number=form.phone_number.data).first()
            existing_username = CommunityMember.query.filter_by(username=form.username.data).first()
            
            if existing_email:
                flash('Bu e-posta adresi zaten kayıtlı.', 'error')
                return render_template('community_register.html', title='Topluluk Üyeliği Başvuru', form=form)
            
            if existing_student_id:
                flash('Bu öğrenci numarası zaten kayıtlı.', 'error')
                return render_template('community_register.html', title='Topluluk Üyeliği Başvuru', form=form)
                
            if existing_phone:
                flash('Bu telefon numarası zaten kayıtlı.', 'error')
                return render_template('community_register.html', title='Topluluk Üyeliği Başvuru', form=form)
                
            if existing_username:
                flash('Bu kullanıcı adı zaten kayıtlı.', 'error')
                return render_template('community_register.html', title='Topluluk Üyeliği Başvuru', form=form)
            
            # Şifreyi hashle
            hashed_password = generate_password_hash(form.password.data, method='pbkdf2:sha256')
            
            community_member = CommunityMember(
                username=form.username.data.strip().lower(),
                password=hashed_password,
                email=form.email.data.lower().strip(),
                name=form.name.data.strip().title(),
                surname=form.surname.data.strip().title(),
                place_of_birth=form.place_of_birth.data.strip().title(),
                date_of_birth=form.date_of_birth.data,
                current_residence=form.current_residence.data.strip().title(),
                student_id=form.student_id.data.strip(),
                phone_number=form.phone_number.data.strip(),
                student_class=form.student_class.data.strip(),
                faculty=form.faculty.data.strip().title(),
                department=form.department.data.strip().title(),
                preferred_units=preferred_units
            )
            db.session.add(community_member)
            db.session.commit()
            
            flash('Topluluk üyeliği başvurunuz başarıyla alındı! Başvurunuz onaylandığında sisteme giriş yapabileceksiniz.', 'success')
            return redirect(url_for('landing'))
            
        except Exception as e:
            db.session.rollback()
            flash('Başvuru gönderilirken bir hata oluştu. Lütfen tekrar deneyin.', 'error')
            app.logger.error(f"Community registration error: {str(e)}")
            
    elif request.method == 'POST':
        # Form validasyon hatalarını kullanıcıya göster
        for field, errors in form.errors.items():
            for error in errors:
                flash(f'{getattr(form, field).label.text}: {error}', 'error')
    
    return render_template('community_register.html', title='Topluluk Üyeliği Başvuru', form=form)

@app.route('/')
@simple_cache.cached(timeout=60, key_prefix='route_')  # 1 dakika cache
def landing():
    return render_template('landing.html')

@app.route('/fansub')
def index():
    # Kullanıcı bazlı cache key oluştur
    cache_key = f'fansub_index_{current_user.id if current_user.is_authenticated else "anonymous"}'
    
    # Cache'den veri almaya çalış
    cached_data = simple_cache.get(cache_key)
    if cached_data:
        return render_template('fansub_index.html', **cached_data)
    
    # Optimized queries with joins
    hero_animes = (Anime.query
                  .join(anime_genres)
                  .join(Genre)
                  .filter(Genre.name == 'Hero Section')
                  .limit(6)
                  .all())
    
    editor_picks = (Anime.query
                   .join(anime_genres)
                   .join(Genre)
                   .filter(Genre.name == 'Editörün Seçimi')
                   .limit(6)
                   .all())
    
    # Yeni eklenenler (son eklenen animeler)
    latest_animes = (Anime.query
                    .order_by(Anime.id.desc())
                    .limit(8)
                    .all())
    
    # Random animeler
    random_animes = (Anime.query
                    .order_by(db.func.random())
                    .limit(8)
                    .all())
    
    personalized_recs = []
    if current_user.is_authenticated:
        user_ratings = db.session.query(Rating.anime_id, db.func.count(Rating.anime_id)).filter_by(user_id=current_user.id).group_by(Rating.anime_id).limit(3).all()
        if user_ratings:
            anime_ids = [rating[0] for rating in user_ratings]
            genre_counts = db.session.query(Genre.id, db.func.count(Genre.id)).join(anime_genres).filter(anime_genres.c.anime_id.in_(anime_ids)).group_by(Genre.id).order_by(db.func.count(Genre.id).desc()).limit(3).all()
            top_genre_ids = [genre[0] for genre in genre_counts]
            personalized_recs = (Anime.query
                                .join(anime_genres)
                                .filter(anime_genres.c.genre_id.in_(top_genre_ids))
                                .limit(6)
                                .all())
    
    # Veriyi cache'e kaydet (2 dakika)
    template_data = {
        'hero_animes': hero_animes, 
        'editor_picks': editor_picks, 
        'latest_animes': latest_animes,
        'random_animes': random_animes,
        'personalized_recs': personalized_recs
    }
    simple_cache.set(cache_key, template_data, timeout=120)
    
    return render_template('fansub_index.html', **template_data)

@app.route('/community')
def community_hub():
    info = CommunityInfo.query.first()
    latest_news = News.query.order_by(News.timestamp.desc()).limit(5).all()
    latest_threads = ForumThread.query.order_by(ForumThread.timestamp.desc()).limit(5).all()
    return render_template('community_hub.html', info=info, latest_news=latest_news, latest_threads=latest_threads)

@app.route('/humat')
def humat_index():
    return redirect(url_for('community_hub'))

@app.route('/animes', methods=['GET', 'POST'])
def animes():
    form = AnimeSearchForm(request.values)
    genres_query = Genre.query
    is_admin = current_user.is_authenticated and (current_user.can_add_user or current_user.can_edit or current_user.can_delete)
    if not is_admin:
        genres_query = genres_query.filter(Genre.name.notin_(SPECIAL_GENRES))
    form.genre.choices = [('', 'Tüm Türler')] + [(str(g.id), g.name) for g in genres_query.order_by('name').all()]
    
    # Optimized query with eager loading
    query = Anime.query.options(db.joinedload(Anime.genres))
    
    if form.validate_on_submit() or request.method == 'GET':
        search_query = sanitize_input(form.query.data or request.args.get('query'))
        if search_query:
            safe_filter = safe_search_filter(search_query, Anime)
            if safe_filter is not None:
                query = query.filter(safe_filter)
    
    if form.genre.data:
        query = query.join(anime_genres).filter(anime_genres.c.genre_id == int(form.genre.data))
    
    if form.release_year.data:
        query = query.filter(Anime.release_year == form.release_year.data)
    
    if form.anime_type.data:
        query = query.filter(Anime.anime_type == form.anime_type.data)
    
    # Optimized sorting
    sort_option = form.sort_by.data or 'name_asc'
    sort_mapping = {
        'name_asc': Anime.name.asc(),
        'name_desc': Anime.name.desc(),
        'rating_desc': Anime.average_rating.desc(),
        'year_desc': Anime.release_year.desc()
    }
    query = query.order_by(sort_mapping.get(sort_option, Anime.name.asc()))
    
    page = request.args.get('page', 1, type=int)
    animes = query.paginate(
        page=page, 
        per_page=18, 
        error_out=False,
        max_per_page=50
    )
    return render_template('all_animes.html', animes=animes, form=form)

@app.route('/episode/<int:episode_id>')
def episode(episode_id):
    episode = Episode.query.get_or_404(episode_id)
    sources = episode.sources.split(',')
    user_genres = session.get('user_genres', {})
    for genre in episode.anime.genres:
        if genre.name not in SPECIAL_GENRES:
            user_genres[str(genre.id)] = user_genres.get(str(genre.id), 0) + 1
    session['user_genres'] = user_genres
    return render_template('episode.html', episode=episode, sources=sources)

@app.route('/admin/genres', methods=['GET', 'POST'])
@login_required
@admin_required
def manage_genres():
    form = GenreForm()
    if form.validate_on_submit():
        new_genre_name = form.name.data
        if new_genre_name in SPECIAL_GENRES:
            flash(f'"{new_genre_name}" türü özel bir türdür ve eklenemez.', 'danger')
        elif Genre.query.filter_by(name=new_genre_name).first():
            flash(f'"{new_genre_name}" türü zaten mevcut.', 'warning')
        else:
            new_genre = Genre(name=new_genre_name)
            db.session.add(new_genre)
            db.session.commit()
            flash(f'"{new_genre_name}" türü eklendi.', 'success')
        return redirect(url_for('manage_genres'))
    genres = Genre.query.all()
    return render_template('manage_genres.html', form=form, genres=genres, special_genres=SPECIAL_GENRES)

@app.route('/admin/delete_genre/<int:genre_id>', methods=['POST'])
@login_required
@admin_required
def delete_genre(genre_id):
    validate_csrf(request.form.get('csrf_token'))
    genre = Genre.query.get_or_404(genre_id)
    if genre.name in SPECIAL_GENRES:
        flash(f'"{genre.name}" türü silinemez.', 'danger')
    else:
        db.session.delete(genre)
        db.session.commit()
        flash(f'"{genre.name}" türü silindi.', 'success')
    return redirect(url_for('manage_genres'))

@app.route('/add_anime', methods=['GET', 'POST'])
@login_required
@admin_required
def add_anime():
    form = AnimeForm()
    if form.validate_on_submit():
        new_anime = Anime(name=form.name.data, description=form.description.data, cover_image=form.cover_image.data, release_year=form.release_year.data, status=form.status.data, anime_type=form.anime_type.data)
        if form.mal_score.data:
            new_anime.mal_score = float(form.mal_score.data)
        if form.mal_url.data:
            new_anime.mal_url = form.mal_url.data
        if not new_anime.mal_score:
            update_mal_data(new_anime)
        for genre in form.genres.data:
            new_anime.genres.append(genre)
        db.session.add(new_anime)
        db.session.commit()
        log_action('add', f'Anime "{new_anime.name}" eklendi.')
        return redirect(url_for('admin'))
    return render_template('add_anime.html', form=form)

@app.route('/edit_anime/<int:anime_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_anime(anime_id):
    anime = Anime.query.get_or_404(anime_id)
    form = AnimeForm(obj=anime)
    if form.validate_on_submit():
        anime.name = form.name.data
        anime.description = form.description.data
        anime.cover_image = form.cover_image.data
        anime.release_year = form.release_year.data
        anime.status = form.status.data
        anime.anime_type = form.anime_type.data
        if form.mal_score.data:
            anime.mal_score = float(form.mal_score.data)
        if form.mal_url.data:
            anime.mal_url = form.mal_url.data
        if not anime.mal_score:
            update_mal_data(anime)
        db.session.commit()
        log_action('update', f'Anime "{anime.name}" düzenlendi.')
        return redirect(url_for('admin'))
    assigned_genres = anime.genres
    unassigned_genres = [genre for genre in Genre.query.all() if genre not in assigned_genres]
    return render_template('edit_anime.html', form=form, anime=anime, assigned_genres=assigned_genres, unassigned_genres=unassigned_genres)

@app.route('/copyright')
def copyright():
    return render_template('copyright.html')

@app.route('/anime/<int:anime_id>')
def anime(anime_id):
    anime = Anime.query.get_or_404(anime_id)
    user_rating = None
    is_in_watchlist = False
    if current_user.is_authenticated:
        user_rating = Rating.query.filter_by(user_id=current_user.id, anime_id=anime.id).first()
        is_in_watchlist = anime in current_user.watchlist_animes
    return render_template('anime.html', anime=anime, user_rating=user_rating, is_in_watchlist=is_in_watchlist)

@app.route('/login', methods=['GET', 'POST'])
@limiter.limit("5 per minute")  # Brute force saldırı koruması
def login():
    # Zaten giriş yapmış kullanıcıyı anasayfaya yönlendir
    if current_user.is_authenticated:
        return redirect(url_for('landing'))
        
    form = LoginForm()
    if form.validate_on_submit():
        # Input güvenliği - Sanitize edilmiş veri
        username_input = sanitize_input(form.username.data.lower().strip())
        
        # Kullanıcı adı veya e-posta ile giriş - Güvenli sorgu
        user = User.query.filter(
            (User.username == username_input) | 
            (User.email == username_input)
        ).first()
        
        if user and check_password_hash(user.password, form.password.data):
            login_user(user, remember=True)
            
            # Session'ı güçlendir
            session.permanent = True
            
            # 🔒 GÜVENLİK: Başarılı login logu
            app.logger.info(f'✅ Başarılı login: Kullanıcı={user.username}, IP={request.remote_addr}, UserAgent={request.headers.get("User-Agent", "Unknown")}')
            log_action('successful_login', f'Başarılı giriş - IP: {request.remote_addr}')
            
            # Eski cache'leri temizle
            simple_cache.delete('fansub_index_anonymous')
            simple_cache.delete(f'fansub_index_{user.id}')
            
            # Flash mesaj
            flash(f'Hoş geldin, {user.username}!', 'success')
            
            # Bir sonraki sayfaya yönlendir
            next_page = request.args.get('next')
            if next_page and next_page.startswith('/'):
                return redirect(next_page)
            return redirect(url_for('landing'))
        else:
            # 🔒 GÜVENLİK: Başarısız login logu
            app.logger.warning(f'❌ Başarısız login denemesi: Username={username_input}, IP={request.remote_addr}, UserAgent={request.headers.get("User-Agent", "Unknown")}')
            
            # Rate limiting için ek gecikme
            import time
            time.sleep(1)  # Brute force yavaşlatması
            
            flash('Giriş başarısız. Lütfen kullanıcı adınızı/e-postanızı ve şifrenizi kontrol edin.', 'danger')
    
    return render_template('login.html', form=form)

@app.route('/admin', methods=['GET', 'POST'])
@login_required
@admin_required
def admin():
    animes = Anime.query.all()
    return render_template('admin.html', animes=animes)

@app.route('/add_episode/<int:anime_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def add_episode(anime_id):
    form = EpisodeForm()
    if form.validate_on_submit():
        episode = Episode(number=form.number.data, sources=form.sources.data, anime_id=anime_id)
        db.session.add(episode)
        db.session.commit()
        anime = db.session.get(Anime, anime_id)
        # Bildirim gönder
        for user in anime.watchlisted_by:
            notification = Notification(message=f"{anime.name} animesinin {episode.number}. bölümü yayınlandı!", user_id=user.id, anime_id=anime.id)
            db.session.add(notification)
        db.session.commit()

        log_action('add', f'Anime "{anime.name}" için bölüm {form.number.data} eklendi.')
        flash('Bölüm başarıyla eklendi.', 'success')
        return redirect(url_for('anime', anime_id=anime_id))
    return render_template('add_episode.html', form=form)

@app.route('/delete_anime/<int:anime_id>', methods=['POST'])
@login_required
@admin_required
def delete_anime(anime_id):
    validate_csrf(request.form.get('csrf_token'))
    if not current_user.can_delete:
        flash('Silme yetkiniz yok!', 'danger')
        return redirect(url_for('admin'))
    anime = Anime.query.get_or_404(anime_id)
    if anime:
        anime_name = anime.name
        Episode.query.filter_by(anime_id=anime_id).delete()
        db.session.delete(anime)
        db.session.commit()
        log_action('delete', f'Anime "{anime_name}" silindi.')
        flash('Anime ve ilgili bölümler başarıyla silindi.', 'success')
    return redirect(url_for('admin'))

@app.route('/delete_episode/<int:episode_id>', methods=['POST'])
@login_required
@admin_required
def delete_episode(episode_id):
    validate_csrf(request.form.get('csrf_token'))
    if not current_user.can_delete:
        flash('Bu işlemi gerçekleştirme yetkiniz yok.', 'danger')
        return redirect(url_for('anime', anime_id=Episode.query.get_or_404(episode_id).anime_id))
    episode = Episode.query.get_or_404(episode_id)
    anime_id = episode.anime_id
    db.session.delete(episode)
    db.session.commit()
    log_action('delete', f'Bölüm {episode.number} silindi. Anime ID: {anime_id}')
    flash('Bölüm başarıyla silindi.', 'success')
    return redirect(url_for('anime', anime_id=anime_id))

@app.route('/edit_episode/<int:episode_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_episode(episode_id):
    if not current_user.can_edit:
        flash('Bu işlemi gerçekleştirme yetkiniz yok.', 'danger')
        return redirect(url_for('anime', anime_id=Episode.query.get_or_404(episode_id).anime_id))
    episode = Episode.query.get_or_404(episode_id)
    form = EpisodeForm()
    if form.validate_on_submit():
        old_number = episode.number
        old_sources = episode.sources
        episode.number = form.number.data
        episode.sources = form.sources.data
        db.session.commit()
        log_action('edit', f'Bölüm {old_number} güncellendi. Yeni numara: {form.number.data}, Eski kaynaklar: {old_sources}, Yeni kaynaklar: {form.sources.data}')
        flash('Bölüm başarıyla güncellendi.', 'success')
        return redirect(url_for('anime', anime_id=episode.anime_id))
    form.number.data = episode.number
    form.sources.data = episode.sources
    return render_template('edit_episode.html', form=form, episode=episode)

@app.route('/add_user', methods=['GET', 'POST'])
@login_required
@admin_required
def add_user():
    if not current_user.can_add_user:
        flash('Yetkiniz yok!', 'danger')
        return redirect(url_for('admin'))
    form = UserForm()
    if form.validate_on_submit():
        hashed_password = generate_password_hash(form.password.data, method='pbkdf2:sha256')
        new_user = User(
            username=form.username.data, 
            password=hashed_password, 
            can_delete=form.can_delete.data, 
            can_edit=form.can_edit.data, 
            can_add_user=form.can_add_user.data,
            is_admin=form.is_admin.data,
            is_community_member=form.is_community_member.data
        )
        db.session.add(new_user)
        db.session.commit()
        log_action('add_user', f'Yeni kullanıcı eklendi: {form.username.data}, Silme: {form.can_delete.data}, Düzenleme: {form.can_edit.data}, Kullanıcı ekleme: {form.can_add_user.data}, Admin: {form.is_admin.data}, Topluluk üyesi: {form.is_community_member.data}')
        flash('Kullanıcı başarıyla eklendi!', 'success')
        return redirect(url_for('admin'))
    return render_template('add_user.html', form=form)

@app.route('/logs')
@login_required
@admin_required
def view_logs():
    if not current_user.can_add_user:
        flash('Bu sayfayı görüntüleme yetkiniz yok.', 'danger')
        return redirect(url_for('index'))
    page = request.args.get('page', 1, type=int)
    logs = Log.query.order_by(Log.timestamp.desc()).paginate(page=page, per_page=10)
    return render_template('log.html', logs=logs)

@app.route('/users')
@login_required
@admin_required
def users():
    if not current_user.can_add_user:
        flash('Bu sayfayı görüntüleme yetkiniz yok.', 'danger')
        return redirect(url_for('index'))
    users = User.query.all()
    return render_template('users.html', users=users)

@app.route('/edit_user/<int:user_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_user(user_id):
    if not current_user.can_add_user:
        flash('Bu işlemi gerçekleştirme yetkiniz yok.', 'danger')
        return redirect(url_for('admin'))
    user = User.query.get_or_404(user_id)
    form = EditUserForm(obj=user)
    if form.validate_on_submit():
        user.username = form.username.data
        user.can_delete = form.can_delete.data
        user.can_edit = form.can_edit.data
        user.can_add_user = form.can_add_user.data
        user.is_admin = form.is_admin.data
        user.is_community_member = form.is_community_member.data
        db.session.commit()
        log_action('edit_user', f'Kullanıcı "{user.username}" güncellendi. Admin: {user.is_admin}, Topluluk üyesi: {user.is_community_member}')
        flash('Kullanıcı başarıyla güncellendi.', 'success')
        return redirect(url_for('admin'))
    return render_template('edit_user.html', form=form, user=user)

@app.route('/api/watchlist/<int:anime_id>', methods=['POST'])
@login_required
@csrf.exempt  # API endpoint için CSRF exempt
def toggle_watchlist(anime_id):
    anime = Anime.query.get_or_404(anime_id)
    if anime in current_user.watchlist_animes:
        current_user.watchlist_animes.remove(anime)
        status = 'removed'
    else:
        current_user.watchlist_animes.append(anime)
        status = 'added'
    db.session.commit()
    return jsonify({'status': status})

@app.route('/api/rate/<int:anime_id>', methods=['POST'])
@login_required
@csrf.exempt  # API endpoint için CSRF exempt
def rate_anime(anime_id):
    anime = Anime.query.get_or_404(anime_id)
    score = request.json.get('score')
    if not score or not 1 <= score <= 5:
        return jsonify({'status': 'error', 'message': 'Invalid score'}), 400
    rating = Rating.query.filter_by(user_id=current_user.id, anime_id=anime.id).first()
    if rating:
        rating.score = score
    else:
        rating = Rating(user_id=current_user.id, anime_id=anime.id, score=score)
        db.session.add(rating)
    total_ratings = db.session.query(db.func.sum(Rating.score)).filter(Rating.anime_id == anime.id).scalar()
    count_ratings = db.session.query(db.func.count(Rating.id)).filter(Rating.anime_id == anime.id).scalar()
    anime.average_rating = total_ratings / count_ratings
    anime.rating_count = count_ratings
    db.session.commit()
    return jsonify({'status': 'success', 'new_average': anime.average_rating, 'rating_count': anime.rating_count})

@app.route('/profile')
@login_required
def profile():
    watchlist = current_user.watchlist_animes
    user_ratings = Rating.query.filter_by(user_id=current_user.id).order_by(Rating.score.desc()).all()
    return render_template('profile.html', watchlist=watchlist, user_ratings=user_ratings)

@app.route('/api/anime/<int:anime_id>/genre/add/<int:genre_id>', methods=['POST'])
@login_required
@admin_required
@csrf.exempt  # API endpoint için CSRF exempt
def add_genre_to_anime(anime_id, genre_id):
    anime = Anime.query.get_or_404(anime_id)
    genre = Genre.query.get_or_404(genre_id)
    if genre not in anime.genres:
        anime.genres.append(genre)
        db.session.commit()
    return jsonify({'status': 'success'})

@app.route('/api/anime/<int:anime_id>/genre/remove/<int:genre_id>', methods=['POST'])
@login_required
@admin_required
@csrf.exempt  # API endpoint için CSRF exempt
def remove_genre_from_anime(anime_id, genre_id):
    anime = Anime.query.get_or_404(anime_id)
    genre = Genre.query.get_or_404(genre_id)
    if genre in anime.genres:
        anime.genres.remove(genre)
        db.session.commit()
    return jsonify({'status': 'success'})

@app.route('/delete_user/<int:user_id>', methods=['POST'])
@login_required
@admin_required
def delete_user(user_id):
    validate_csrf(request.form.get('csrf_token'))
    user = User.query.get_or_404(user_id)
    if not current_user.can_delete:
        flash('Bu sayfayı görüntüleme yetkiniz yok.', 'danger')
        return redirect(url_for('index'))
    logs = Log.query.filter_by(user_id=user_id).all()
    for log in logs:
        log.user_id = None
    db.session.commit()
    db.session.delete(user)
    db.session.commit()
    flash('Kullanıcı başarıyla silindi!', 'success')
    return redirect(url_for('users'))

@app.route('/logout')
def logout():
    if current_user.is_authenticated:
        username = current_user.username
        user_id = current_user.id
        
        # Kullanıcı çıkış yap
        logout_user()
        
        # Session'ı tamamen temizle
        session.clear()
        
        # Kullanıcı bazlı cache'i temizle
        user_cache_key = f'fansub_index_{user_id}'
        simple_cache.delete(user_cache_key)
        
        # Anonymous cache'i de temizle
        simple_cache.delete('fansub_index_anonymous')
        
        flash(f'Güvenle çıkış yaptınız, {username}.', 'info')
    else:
        flash('Zaten çıkış yapmışsınız.', 'info')
    
    return redirect(url_for('login'))

@app.route('/support')
def support():
    return render_template('support.html')

@app.route('/api/notifications')
@login_required
def get_notifications():
    notifications = Notification.query.filter_by(user_id=current_user.id).order_by(Notification.timestamp.desc()).all()
    return jsonify([{'id': n.id, 'message': n.message, 'timestamp': n.timestamp.isoformat(), 'is_read': n.is_read, 'anime_id': n.anime_id} for n in notifications])

@app.route('/api/notifications/mark_read', methods=['POST'])
@login_required
@csrf.exempt  # API endpoint için CSRF exempt
def mark_notifications_as_read():
    Notification.query.filter_by(user_id=current_user.id, is_read=False).update({'is_read': True})
    db.session.commit()
    return jsonify({'status': 'success'})

# News and Events Routes
@app.route('/news')
def news():
    news_list = News.query.order_by(News.is_pinned.desc(), News.timestamp.desc()).all()
    return render_template('news.html', news_list=news_list)

@app.route('/news/<int:news_id>')
def view_news(news_id):
    news_item = News.query.get_or_404(news_id)
    return render_template('news_article.html', news_item=news_item)

@app.route('/admin/news', methods=['GET', 'POST'])
@login_required
@admin_required
def manage_news():
    form = NewsForm()
    if form.validate_on_submit():
        new_news = News(title=form.title.data, content=form.content.data, image_url=form.image_url.data, user_id=current_user.id)
        db.session.add(new_news)
        db.session.commit()
        flash('Haber başarıyla eklendi.', 'success')
        return redirect(url_for('manage_news'))
    all_news = News.query.order_by(News.is_pinned.desc(), News.timestamp.desc()).all()
    return render_template('manage_news.html', form=form, all_news=all_news)

@app.route('/admin/news/edit/<int:news_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_news(news_id):
    news_item = News.query.get_or_404(news_id)
    form = NewsForm(obj=news_item)
    if form.validate_on_submit():
        news_item.title = form.title.data
        news_item.content = form.content.data
        news_item.image_url = form.image_url.data
        news_item.is_pinned = form.is_pinned.data
        db.session.commit()
        flash('Haber başarıyla güncellendi.', 'success')
        return redirect(url_for('manage_news'))
    return render_template('edit_news.html', form=form, news_item=news_item)

@app.route('/admin/news/delete/<int:news_id>', methods=['POST'])
@login_required
@admin_required
def delete_news(news_id):
    news_item = News.query.get_or_404(news_id)
    db.session.delete(news_item)
    db.session.commit()
    flash('Haber silindi.', 'success')
    return redirect(url_for('manage_news'))

@app.route('/admin/news/pin/<int:news_id>', methods=['POST'])
@login_required
@admin_required
def pin_news(news_id):
    news_item = News.query.get_or_404(news_id)
    news_item.is_pinned = not news_item.is_pinned
    db.session.commit()
    flash('Haberin sabitleme durumu değiştirildi.', 'success')
    return redirect(url_for('manage_news'))

@app.route('/events')
@login_required
@calendar_access_required
def events():
    return render_template('events.html')

@app.route('/admin/events', methods=['GET', 'POST'])
@login_required
@admin_required
def manage_events():
    form = EventForm()
    if form.validate_on_submit():
        start_time = datetime.strptime(form.start_time.data, '%Y-%m-%dT%H:%M')
        end_time = datetime.strptime(form.end_time.data, '%Y-%m-%dT%H:%M') if form.end_time.data else None
        new_event = Event(title=form.title.data, 
                            description=form.description.data, 
                            start_time=start_time, 
                            end_time=end_time, 
                            user_id=current_user.id)
        db.session.add(new_event)
        db.session.commit()
        flash('Etkinlik başarıyla eklendi.', 'success')
        return redirect(url_for('manage_events'))
    all_events = Event.query.order_by(Event.start_time.desc()).all()
    return render_template('manage_events.html', form=form, all_events=all_events)

@app.route('/admin/events/edit/<int:event_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_event(event_id):
    event_item = Event.query.get_or_404(event_id)
    form = EventForm(obj=event_item)
    if form.validate_on_submit():
        event_item.title = form.title.data
        event_item.description = form.description.data
        event_item.start_time = datetime.strptime(form.start_time.data, '%Y-%m-%dT%H:%M')
        event_item.end_time = datetime.strptime(form.end_time.data, '%Y-%m-%dT%H:%M') if form.end_time.data else None
        db.session.commit()
        flash('Etkinlik başarıyla güncellendi.', 'success')
        return redirect(url_for('manage_events'))
    return render_template('edit_event.html', form=form, event_item=event_item)

@app.route('/admin/events/delete/<int:event_id>', methods=['POST'])
@login_required
@admin_required
def delete_event(event_id):
    validate_csrf(request.form.get('csrf_token'))
    event_item = Event.query.get_or_404(event_id)
    db.session.delete(event_item)
    db.session.commit()
    flash('Etkinlik silindi.', 'success')
    return redirect(url_for('manage_events'))

@app.route('/api/events')
def api_events():
    events = Event.query.all()
    event_list = []
    for event in events:
        event_list.append({
            'title': event.title,
            'start': event.start_time.isoformat(),
            'end': event.end_time.isoformat() if event.end_time else None,
            'description': event.description,
            'url': url_for('edit_event', event_id=event.id)
        })
    return jsonify(event_list)

@app.route('/admin/community_members', methods=['GET', 'POST'])
@login_required
@strict_admin_required
def manage_community_members():
    form = CommunityMemberSearchForm()
    
    # Temel query'ler - User relationship'ini de yükle
    pending_query = CommunityMember.query.options(db.joinedload(CommunityMember.user)).filter_by(is_approved=False)
    approved_query = CommunityMember.query.options(db.joinedload(CommunityMember.user)).filter_by(is_approved=True)
    
    # Arama parametreleri - güvenli hale getir
    search_query = sanitize_input(request.args.get('query', '').strip())
    status_filter = request.args.get('status', '')
    
    # Form'dan değerleri al
    if form.validate_on_submit():
        search_query = sanitize_input(form.query.data.strip() if form.query.data else '')
        status_filter = form.status.data
    
    # Form alanlarını doldur
    form.query.data = search_query
    form.status.data = status_filter
    
    # Arama filtreleri uygula - güvenli şekilde
    if search_query:
        search_filter = safe_search_filter(search_query, CommunityMember)
        if search_filter is not None:
            pending_query = pending_query.filter(search_filter)
            approved_query = approved_query.filter(search_filter)
    
    # Durum filtresi uygula
    if status_filter == 'pending':
        pending_members = pending_query.order_by(CommunityMember.registration_date.desc()).all()
        approved_members = []
    elif status_filter == 'approved':
        pending_members = []
        approved_members = approved_query.order_by(CommunityMember.registration_date.desc()).all()
    else:
        pending_members = pending_query.order_by(CommunityMember.registration_date.desc()).all()
        approved_members = approved_query.order_by(CommunityMember.registration_date.desc()).all()
    
    # Debug: Template'e gönderilen verileri logla
    app.logger.info(f"Community Members Page - Search: '{search_query}', Status: '{status_filter}'")
    app.logger.info(f"Pending members count: {len(pending_members)}")
    app.logger.info(f"Approved members count: {len(approved_members)}")
    
    for member in approved_members:
        app.logger.info(f"Approved member: ID={member.id}, Name={member.name} {member.surname}, Username={member.username}, User={member.user.username if member.user else 'None'}")
    
    return render_template('admin_community_members.html', 
                         pending_members=pending_members, 
                         approved_members=approved_members,
                         form=form,
                         search_query=search_query,
                         status_filter=status_filter)

@app.route('/admin/community_members/export_excel')
@login_required
@strict_admin_required
@limiter.limit("5 per minute")  # Rate limiting for export
def export_community_members_excel():
    """Topluluk üyelerinin verilerini güvenli bir şekilde Excel formatında indir"""
    try:
        # IP ve kullanıcı bilgilerini logla (güvenlik için)
        client_ip = get_remote_address()
        app.logger.info(f"Excel export requested by user: {current_user.username} (ID: {current_user.id}) from IP: {client_ip}")
        
        # Critical action logu
        log_action('excel_export', f'Topluluk üyeleri Excel export - Admin: {current_user.username}, IP: {client_ip}')
        
        # IP-based güvenlik (isteğe bağlı - environment variable ile kontrol)
        allowed_ips = os.environ.get('ADMIN_ALLOWED_IPS', '').split(',')
        if allowed_ips and allowed_ips != [''] and client_ip not in allowed_ips:
            app.logger.warning(f"🚨 Excel export denied for IP: {client_ip}, User: {current_user.username}")
            flash('Bu işlem için IP adresiniz yetkili değil.', 'error')
            return redirect(url_for('manage_community_members'))
        
        # Workbook oluştur
        wb = Workbook()
        
        # Onaylı üyeler sayfası
        ws_approved = wb.active
        ws_approved.title = "Onaylı Üyeler"
        
        # Bekleyen üyeler sayfası
        ws_pending = wb.create_sheet("Bekleyen Başvurular")
        
        # Başlık stilleri
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Onaylı üyeler için başlıklar
        approved_headers = [
            "ID", "Ad", "Soyad", "Kullanıcı Adı", "E-posta", "Öğrenci No",
            "Telefon", "Doğum Yeri", "Doğum Tarihi", "İkamet", "Sınıf",
            "Fakülte", "Bölüm", "Tercih Ettiği Birimler", "Kayıt Tarihi", "Sistem Kullanıcı Adı"
        ]
        
        # Bekleyen başvurular için başlıklar
        pending_headers = [
            "ID", "Ad", "Soyad", "Kullanıcı Adı", "E-posta", "Öğrenci No",
            "Telefon", "Doğum Yeri", "Doğum Tarihi", "İkamet", "Sınıf",
            "Fakülte", "Bölüm", "Tercih Ettiği Birimler", "Başvuru Tarihi"
        ]
        
        # Onaylı üyeler başlıklarını yaz
        for col_num, header in enumerate(approved_headers, 1):
            cell = ws_approved.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Bekleyen başvurular başlıklarını yaz
        for col_num, header in enumerate(pending_headers, 1):
            cell = ws_pending.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Onaylı üyelerin verilerini al ve yaz
        approved_members = CommunityMember.query.options(db.joinedload(CommunityMember.user)).filter_by(is_approved=True).order_by(CommunityMember.registration_date.desc()).all()
        
        for row_num, member in enumerate(approved_members, 2):
            # Güvenlik: Hassas verileri sanitize et
            ws_approved.cell(row=row_num, column=1, value=member.id)
            ws_approved.cell(row=row_num, column=2, value=sanitize_input(member.name))
            ws_approved.cell(row=row_num, column=3, value=sanitize_input(member.surname))
            ws_approved.cell(row=row_num, column=4, value=sanitize_input(member.username))
            ws_approved.cell(row=row_num, column=5, value=sanitize_input(member.email))
            ws_approved.cell(row=row_num, column=6, value=sanitize_input(member.student_id))
            # Telefon numarasını tam olarak yaz (admin için)
            ws_approved.cell(row=row_num, column=7, value=sanitize_input(member.phone_number))
            ws_approved.cell(row=row_num, column=8, value=sanitize_input(member.place_of_birth))
            ws_approved.cell(row=row_num, column=9, value=member.date_of_birth.strftime('%Y-%m-%d') if member.date_of_birth else '')
            ws_approved.cell(row=row_num, column=10, value=sanitize_input(member.current_residence))
            ws_approved.cell(row=row_num, column=11, value=sanitize_input(member.student_class))
            ws_approved.cell(row=row_num, column=12, value=sanitize_input(member.faculty))
            ws_approved.cell(row=row_num, column=13, value=sanitize_input(member.department))
            ws_approved.cell(row=row_num, column=14, value=sanitize_input(member.preferred_units) if member.preferred_units else '')
            ws_approved.cell(row=row_num, column=15, value=member.registration_date.strftime('%Y-%m-%d %H:%M') if member.registration_date else '')
            ws_approved.cell(row=row_num, column=16, value=member.user.username if member.user else 'N/A')
        
        # Bekleyen başvuruların verilerini al ve yaz
        pending_members = CommunityMember.query.filter_by(is_approved=False).order_by(CommunityMember.registration_date.desc()).all()
        
        for row_num, member in enumerate(pending_members, 2):
            # Güvenlik: Hassas verileri sanitize et
            ws_pending.cell(row=row_num, column=1, value=member.id)
            ws_pending.cell(row=row_num, column=2, value=sanitize_input(member.name))
            ws_pending.cell(row=row_num, column=3, value=sanitize_input(member.surname))
            ws_pending.cell(row=row_num, column=4, value=sanitize_input(member.username))
            ws_pending.cell(row=row_num, column=5, value=sanitize_input(member.email))
            ws_pending.cell(row=row_num, column=6, value=sanitize_input(member.student_id))
            # Telefon numarasını tam olarak yaz (admin için)
            ws_pending.cell(row=row_num, column=7, value=sanitize_input(member.phone_number))
            ws_pending.cell(row=row_num, column=8, value=sanitize_input(member.place_of_birth))
            ws_pending.cell(row=row_num, column=9, value=member.date_of_birth.strftime('%Y-%m-%d') if member.date_of_birth else '')
            ws_pending.cell(row=row_num, column=10, value=sanitize_input(member.current_residence))
            ws_pending.cell(row=row_num, column=11, value=sanitize_input(member.student_class))
            ws_pending.cell(row=row_num, column=12, value=sanitize_input(member.faculty))
            ws_pending.cell(row=row_num, column=13, value=sanitize_input(member.department))
            ws_pending.cell(row=row_num, column=14, value=sanitize_input(member.preferred_units) if member.preferred_units else '')
            ws_pending.cell(row=row_num, column=15, value=member.registration_date.strftime('%Y-%m-%d %H:%M') if member.registration_date else '')
        
        # Kolon genişliklerini ayarla
        for ws in [ws_approved, ws_pending]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Maksimum 50 karakter
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Excel dosyasını memory'de oluştur
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Dosya adını güvenli şekilde oluştur
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"topluluk_uyeleri_{timestamp}.xlsx"
        
        # Güvenlik logu
        app.logger.info(f"Excel export completed: {len(approved_members)} approved members, {len(pending_members)} pending applications exported by {current_user.username}")
        
        return send_file(
            excel_buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        app.logger.error(f"Excel export error: {str(e)}")
        flash('Excel dosyası oluşturulurken bir hata oluştu.', 'error')
        return redirect(url_for('manage_community_members'))

@app.route('/admin/community_members/approve/<int:member_id>', methods=['POST'])
@login_required
@strict_admin_required
def approve_community_member(member_id):
    validate_csrf(request.form.get('csrf_token'))
    member = CommunityMember.query.get_or_404(member_id)
    
    try:
        # Topluluk üyesinin seçtiği kullanıcı adının mevcut User tablosunda olup olmadığını kontrol et
        existing_user = User.query.filter_by(username=member.username).first()
        if existing_user:
            flash(f'Kullanıcı adı "{member.username}" zaten sistemde mevcut. Üyeye farklı bir kullanıcı adı seçmesini söyleyin.', 'error')
            return redirect(url_for('manage_community_members'))
        
        # Topluluk üyesinin bilgilerini kullanarak User kaydı oluştur
        new_user = User(
            username=member.username,
            email=member.email,
            password=member.password,  # Zaten hashlenmiş şifre
            is_community_member=True
        )
        db.session.add(new_user)
        db.session.flush()  # ID'yi almak için
        
        # Topluluk üyesini güncelle
        member.is_approved = True
        member.user_id = new_user.id
        
        db.session.commit()
        
        flash(f'{member.name} {member.surname} onaylandı. Kullanıcı adı: {member.username} ile sisteme giriş yapabilir.', 'success')
        
        # Debug log ekle
        log_action('community_member_approval', f'Üye onaylandı: {member.name} {member.surname} ({member.username}) - User ID: {new_user.id}, Community Member ID: {member.id}')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Onay işlemi sırasında hata oluştu: {str(e)}', 'error')
        app.logger.error(f"Community member approval error: {str(e)}")
    
    return redirect(url_for('manage_community_members'))

@app.route('/admin/community_members/reject/<int:member_id>', methods=['POST'])
@login_required
@strict_admin_required
def reject_community_member(member_id):
    validate_csrf(request.form.get('csrf_token'))
    member = CommunityMember.query.get_or_404(member_id)
    db.session.delete(member)
    db.session.commit()
    flash(f'{member.name} {member.surname} başvurusu reddedildi.', 'warning')
    return redirect(url_for('manage_community_members'))

@app.route('/admin/community_members/delete/<int:member_id>', methods=['POST'])
@login_required
@strict_admin_required
@delete_permission_required
def delete_community_member(member_id):
    validate_csrf(request.form.get('csrf_token'))
    if not current_user.can_delete:
        flash('Bu işlemi gerçekleştirme yetkiniz yok.', 'danger')
        return redirect(url_for('manage_community_members'))
    
    member = CommunityMember.query.get_or_404(member_id)
    
    try:
        # Debug: Silme öncesi durum
        app.logger.info(f"Deleting community member: ID={member.id}, Name={member.name} {member.surname}, Username={member.username}, Approved={member.is_approved}, User_ID={member.user_id}")
        
        # Eğer topluluk üyesinin bir kullanıcı hesabı varsa, onu da sil
        if member.user_id:
            user = db.session.get(User, member.user_id)
            if user:
                # Kullanıcıya ait logları temizle (user_id'yi null yap)
                logs = Log.query.filter_by(user_id=user.id).all()
                for log in logs:
                    log.user_id = None
                
                # Kullanıcıyı sil
                app.logger.info(f"Also deleting associated user: ID={user.id}, Username={user.username}")
                db.session.delete(user)
        
        # Topluluk üyesini sil
        member_name = f"{member.name} {member.surname}"
        member_username = member.username
        db.session.delete(member)
        db.session.commit()
        
        flash(f'Topluluk üyesi {member_name} ({member_username}) başarıyla silindi.', 'success')
        log_action('delete_community_member', f'Topluluk üyesi {member_name} ({member_username}) silindi.')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Silme işlemi sırasında hata oluştu: {str(e)}', 'error')
        app.logger.error(f"Community member deletion error: {str(e)}")
    
    return redirect(url_for('manage_community_members'))

@app.route('/admin/community_members/view/<int:member_id>')
@login_required
@strict_admin_required
def view_community_member(member_id):
    member = CommunityMember.query.get_or_404(member_id)
    return render_template('view_community_member.html', member=member)

@app.route('/admin/community_members/export')
@login_required
@strict_admin_required
@limiter.limit("2 per minute")  # Export spam koruması
def export_community_members():
    # 🔒 GÜVENLİK: Hassas veriler çıkarıldı, sadece gerekli bilgiler export ediliyor
    
    # Çoklu admin kontrolü
    if not current_user.is_admin:
        log_action('unauthorized_export_attempt', f'Yetkisiz export denemesi: {current_user.username}')
        flash('Bu işlem için admin yetkisi gerekiyor.', 'danger')
        return redirect(url_for('manage_community_members'))
    
    # IP loglaması
    app.logger.warning(f"Community members export requested by {current_user.username} from IP: {request.remote_addr}")
    
    try:
        # Arama parametrelerini al ve güvenli hale getir
        search_query = sanitize_input(request.args.get('query', '').strip())
        status_filter = request.args.get('status', '')
        
        # Excel dosyası oluştur
        wb = Workbook()
        ws = wb.active
        ws.title = "Topluluk Üyeleri"
        
        # Başlık stilleri
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # ✅ TÜM VERİLER: Admin tarafından talep edilen tüm sütunlar
        headers = [
            "ID", "Ad", "Soyad", "Kullanıcı Adı", "E-posta", "Öğrenci No",
            "Telefon", "Doğum Yeri", "Doğum Tarihi", "İkamet", "Sınıf",
            "Fakülte", "Bölüm", "Tercih Ettiği Birimler", "Başvuru Tarihi", "Onay Durumu"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Verileri filtrele ve al
        query = CommunityMember.query
        
        # Arama filtreleri uygula - güvenli şekilde
        if search_query:
            search_filter = safe_search_filter(search_query, CommunityMember)
            if search_filter is not None:
                query = query.filter(search_filter)
        
        # Durum filtresi uygula
        if status_filter == 'pending':
            query = query.filter_by(is_approved=False)
        elif status_filter == 'approved':
            query = query.filter_by(is_approved=True)
        
        members = query.order_by(CommunityMember.registration_date.desc()).all()
        
        for row, member in enumerate(members, 2):
            ws.cell(row=row, column=1, value=member.id)
            ws.cell(row=row, column=2, value=sanitize_input(member.name))
            ws.cell(row=row, column=3, value=sanitize_input(member.surname))
            ws.cell(row=row, column=4, value=sanitize_input(member.username))
            ws.cell(row=row, column=5, value=sanitize_input(member.email))
            ws.cell(row=row, column=6, value=sanitize_input(member.student_id))
            ws.cell(row=row, column=7, value=sanitize_input(member.phone_number))
            ws.cell(row=row, column=8, value=sanitize_input(member.place_of_birth))
            ws.cell(row=row, column=9, value=member.date_of_birth.strftime('%Y-%m-%d') if member.date_of_birth else '')
            ws.cell(row=row, column=10, value=sanitize_input(member.current_residence))
            ws.cell(row=row, column=11, value=sanitize_input(member.student_class))
            ws.cell(row=row, column=12, value=sanitize_input(member.faculty))
            ws.cell(row=row, column=13, value=sanitize_input(member.department))
            ws.cell(row=row, column=14, value=sanitize_input(member.preferred_units) if member.preferred_units else '')
            ws.cell(row=row, column=15, value=member.registration_date.strftime('%Y-%m-%d %H:%M') if member.registration_date else '')
            ws.cell(row=row, column=16, value="Onaylandı" if member.is_approved else "Beklemede")
        
        # Sütun genişliklerini ayarla (16 sütun için)
        column_widths = [8, 12, 12, 15, 25, 12, 15, 15, 12, 15, 10, 20, 25, 20, 18, 12]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
        
        # Dosyayı belleğe kaydet
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Dosya adı oluştur (arama parametrelerine göre)
        filename_parts = ["topluluk_uyeleri"]
        if search_query:
            filename_parts.append(f"arama_{search_query.replace(' ', '_')}")
        if status_filter:
            filename_parts.append(status_filter)
        filename_parts.append(datetime.now().strftime('%Y%m%d_%H%M%S'))
        filename = "_".join(filename_parts) + ".xlsx"
        
        # Log kaydı - Tüm veriler export edildi
        log_message = f'� FULL EXPORT: Topluluk üyeleri listesi (TÜM VERİLER) Excel olarak dışa aktarıldı. Toplam üye: {len(members)}, Kullanıcı: {current_user.username}, IP: {request.remote_addr}'
        if search_query:
            log_message += f', Arama: "{search_query}"'
        if status_filter:
            log_message += f', Durum: {status_filter}'
        log_action('export_community_members', log_message)
        
        return send_file(
            excel_buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        flash(f'Excel dosyası oluşturulurken hata oluştu: {str(e)}', 'error')
        app.logger.error(f"Excel export error: {str(e)}")
        return redirect(url_for('manage_community_members'))

# Cache clearing helper functions
def clear_anime_cache():
    """Clear anime-related cache"""
    # Tüm fansub index cache'lerini temizle
    for key in simple_cache.cache.keys():
        if key.startswith('fansub_index_'):
            simple_cache.delete(key)
    simple_cache.delete('route_landing:()')

def clear_community_cache():
    """Clear community-related cache"""
    simple_cache.delete('route_community_hub:()')

# Add cache clearing to relevant functions
@app.after_request
def after_request(response):
    # Clear cache on data modifications
    if request.method in ['POST', 'PUT', 'DELETE'] and request.endpoint:
        if 'anime' in request.endpoint:
            clear_anime_cache()
        elif 'community' in request.endpoint:
            clear_community_cache()
    return response

if __name__ == '__main__':
    # Production settings
    import os
    debug_mode = os.environ.get('FLASK_DEBUG', 'False').lower() == 'true'
    port = int(os.environ.get('PORT', 5011))  # Port 5007 kullan
    app.run(debug=debug_mode, host='0.0.0.0', port=port, threaded=True)
    
    